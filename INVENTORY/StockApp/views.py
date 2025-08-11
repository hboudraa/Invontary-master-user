import os
from django.conf import settings
import openpyxl
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.contrib import messages
from .forms import ProductForm, EnterForm, DemandClassForm, SortieForm
from .models import *
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.urls import reverse

import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO
from django.db.models import Q

from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import re

# Create your views here.

def home(request):
    return render(request, 'home.html', {})

def products(request):
    categories = Category.objects.all()
    selected_category = request.GET.get('category', '')
    search_query = request.GET.get('search_query', '')

    # جلب جميع المنتجات
    product_list = Product.objects.all().order_by('id')

    # فلترة حسب الفئة إذا تم اختيارها
    if selected_category:
        product_list = product_list.filter(category_id=selected_category)

    # فلترة حسب البحث إذا تم إدخاله
    if search_query:
        product_list = product_list.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # تقسيم الصفحات
    paginator = Paginator(product_list, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # في حالة طلب Ajax
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html = render_to_string('product_list.html', {
            'page_obj': page_obj,
            'categories': categories,
            'selected_category': selected_category,
            'search_query': search_query
        })
        return JsonResponse({'html': html})

    return render(request, 'list_products.html', {
        'categories': categories,
        'page_obj': page_obj,
        'selected_category': selected_category,
        'search_query': search_query
    })
def add_enter(request, id):
    enter = Product.objects.get(pk=id)  # ✅ جلب كائن المنتج الصحيح
    if request.user.is_authenticated:
        if request.method == 'POST':
            form_enter = EnterForm(request.POST, product=enter)  # ✅ تمرير كائن المنتج للنموذج
            if form_enter.is_valid():
                fiche_enter = form_enter.save(commit=False)
                fiche_enter.user = request.user
                fiche_enter.name_fiche = enter  # ✅ تعيين كائن المنتج بدلاً من اسمه
                fiche_enter.save()
                messages.success(request, "✅ Fiche added successfully!")
                return redirect('product:AddStockEnter', id=id)
            else:
                print(form_enter.errors)  # عرض الأخطاء في الكونسول
                messages.error(request, "❌ Error: Please check the form fields.")
        else:
            form_enter = EnterForm(product=enter)  # ✅ تمرير المنتج عند تحميل الصفحة لأول مرة
        return render(request, 'AddStockEnter.html', {'form_enter': form_enter})
    messages.error(request, "❌ You must be logged in to add a product.")
    return redirect('accounts:login')  # توجيه المستخدم إلى صفحة تسجيل الدخول


def fiche_stock_entr(request, pk):
    item = get_object_or_404(Product, pk=pk)
    list_fiche = FicheStockEntr.objects.filter(name_fiche=item).order_by('id')  # 🔹 تصفية السجلات للمنتج المحدد
    paginator = Paginator(list_fiche, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, 'ficheStockEntr.html', {
        'item': item,
        "page_obj": page_obj
        })


def add_sortie(request, id):
    sortie = Product.objects.get(pk=id)
    ini_quantity = sortie.quantity
    print(ini_quantity)
    if request.user.is_authenticated:
        if request.method == 'POST':
            form_sortie = SortieForm(request.POST, product=sortie)
            if form_sortie.is_valid():
                fiche_sortie = form_sortie.save(commit=False)
                fiche_sortie.user = request.user
                fiche_sortie.name_fiche = sortie
                if form_sortie.cleaned_data['quantity'] <= ini_quantity:
                    fiche_sortie.save()
                    messages.success(request, "✅ Fiche added successfully!")
                    return redirect('product:fiche_stock_sort', pk=id)
                else:
                    messages.error(request, '❌ Error: Please check the QUANTITY form fields.')
            else:
                print(form_sortie.errors)
                messages.error(request,"❌ Error: Please check the form fields.")
        else:
            form_sortie = SortieForm(product=sortie)
        return render(request, 'add-stock-sortie.html', {'form_sortie':form_sortie})
    messages.error(request, "❌ You must be logged in to add a product.")
    return redirect('accounts:login')  # توجيه المستخدم إلى صفحة تسجيل الدخول



@login_required
def batch_sortie(request):
    products = Product.objects.all().order_by('id')

    if request.method == "POST":
        destination = request.POST.get("destination", "").strip()
        number_prefix = request.POST.get("number_prefix", "")
        created_ids = []  # لتخزين IDs سجلات FicheStockSortie

        with transaction.atomic():
            for product in products:
                qty_raw = request.POST.get(f"qty_{product.id}", "").strip()
                if not qty_raw:
                    continue
                try:
                    qty = int(qty_raw)
                except ValueError:
                    continue

                if qty <= 0 or qty > product.quantity:
                    continue

                fiche = FicheStockSortie.objects.create(
                    user=request.user,
                    name_fiche=product,
                    number=f"{number_prefix}-{product.id}",
                    destination=destination or "غير محدد",
                    quantity=qty,
                    observation=request.POST.get(f"obs_{product.id}", "").strip()
                )
                created_ids.append(fiche.id)

        if created_ids:
            # نخزن IDs في الـ session
            request.session['last_batch_ids'] = created_ids
            return redirect(reverse('product:batch_pdf'))
        else:
            messages.info(request, "لم يتم إخراج أي منتج.")
            return redirect('product:batch_sortie')

    return render(request, 'batch_sortie.html', {'products': products})


def fiche_stock_sortie(request, pk):
    item = get_object_or_404(Product, pk=pk)
    list_fiche = FicheStockSortie.objects.filter(name_fiche=item).order_by('id')  # 🔹 تصفية السجلات للمنتج المحدد
    paginator = Paginator(list_fiche, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, 'ficheStockSort.html', {
        'item': item,
        "page_obj": page_obj
    })

def demand(request, id):
    product = get_object_or_404(Product, pk=id)

    if not request.user.is_authenticated:
        messages.error(request, "❌ You must be logged in to add a product.")
        return redirect('accounts:login')

    if request.method == 'POST':
        form = DemandClassForm(request.POST, product=product)
        if form.is_valid():
            demand = form.save(commit=False)
            demand.user = request.user
            demand.name_demand = product
            demand.save()
            messages.success(request, "✅ Demande added successfully!")
            return redirect('product:products')
        else:
            print(form.errors)
            messages.error(request, "❌ Error: Please check the form fields.")
    else:
        form = DemandClassForm(product=product)
    if request.user.is_superuser:
        list_demand = DemandClass.objects.filter(name_demand=product).order_by('id')
    else:
        list_demand = DemandClass.objects.filter(name_demand=product, user=request.user).order_by('id')
    paginator = Paginator(list_demand, 10)
    page_number = (request.GET.get("page"))
    page_obj = paginator.get_page(page_number)
    return render(request, 'demander.html', {
        'form_demand': form,
        'demand_id': product,
        'page_obj':page_obj,
    })

@staff_member_required
def all_demands(request):
    list_demand = DemandClass.objects.all().order_by('-date')
    paginator = Paginator(list_demand, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    return render(request, 'all_demands.html', {
        'page_obj': page_obj
    })


@staff_member_required
def update_demand_status(request, demand_id, status):
    demand = get_object_or_404(DemandClass, pk=demand_id)

    if status in ['accepted', 'rejected']:
        demand.status = status
        demand.save()
        messages.success(request, f"✅ Demande {status} successfully.")
    else:
        messages.error(request, "❌ Invalid status.")

    return redirect(request.META.get('HTTP_REFERER', 'product:products'))
    

def add_product(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            form = ProductForm(request.POST)
            if form.is_valid():
                product = form.save(commit=False)
                product.user = request.user
                product.save()
                messages.success(request, "✅ Product added successfully!")
                return redirect('product:products')
            else:
                print(form.errors)  # Debugging: show errors in console
                messages.error(request, "❌ Error: Please check the form fields.")
        else:
            form = ProductForm()
        return render(request, 'add_product.html', {'form': form})

    # Redirect unauthenticated users
    messages.error(request, "❌ You must be logged in to add a product.")
    return redirect('accounts:login')  # Change 'login' to your actual login URL name



# def export_products_to_excel(request):
#     # Create a workbook and add a worksheet
#     wb = openpyxl.Workbook()
#     ws = wb.active
#     ws.title = "Products"
#
#     # Define the column headers
#     headers = ['ID', 'Name', 'Category', 'Quantity', 'Price']
#
#     # Add headers to the worksheet
#     ws.append(headers)
#
#     # Fetch all products from the database
#     products = Product.objects.all()
#
#     # Add product data to the worksheet
#     for product in products:
#         ws.append(
#             [product.id, product.name, product.category.name, product.quantity, product.price])
#
#     # Prepare the response
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=products.xlsx'
#
#     # Save the workbook to the response
#     wb.save(response)
#
#     return response

@login_required
def export_products_to_excel(request):
    """
    Enhanced Excel export with formatting, charts, and comprehensive data
    """
    # Create a workbook with multiple worksheets
    wb = openpyxl.Workbook()

    # Remove default sheet and create custom sheets
    wb.remove(wb.active)

    # Create main products sheet
    ws_products = wb.create_sheet("Products List")

    # Create summary sheet
    ws_summary = wb.create_sheet("Summary Dashboard")

    # Create low stock alerts sheet
    ws_alerts = wb.create_sheet("Stock Alerts")

    # ============= PRODUCTS SHEET =============
    setup_products_sheet(ws_products)

    # ============= SUMMARY SHEET =============
    setup_summary_sheet(ws_summary)

    # ============= ALERTS SHEET =============
    setup_alerts_sheet(ws_alerts)

    # Set the products sheet as active
    wb.active = ws_products

    # Prepare the response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Generate filename with timestamp
    timestamp = timezone.now().strftime("%Y-%m-%d_%H-%M")
    filename = f'products_export_{timestamp}.xlsx'
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Save the workbook to the response
    wb.save(response)

    return response


def setup_products_sheet(ws):
    """Setup the main products sheet with formatting"""
    from .models import Product  # Import here to avoid circular imports

    # Define comprehensive headers
    headers = [
        'ID', 'Nom du Produit', 'Catégorie', 'Description', 'Quantité',
        'Prix (€)', 'Date d\'Expiration', 'Jours Restants', 'Statut Stock',
        'Valeur Total (€)', 'Utilisateur', 'Date Création'
    ]

    # Add title
    ws.merge_cells('A1:L1')
    title_cell = ws['A1']
    title_cell.value = f"RAPPORT D'INVENTAIRE - {timezone.now().strftime('%d/%m/%Y à %H:%M')}"
    title_cell.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add headers in row 3
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='A23B72', end_color='A23B72', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Fetch products with related data
    products = Product.objects.select_related('category', 'user').all()

    # Add product data starting from row 4
    for row_num, product in enumerate(products, 4):
        # Calculate remaining days and stock status
        remaining_days = None
        if product.expiration_date:
            remaining_days = (product.expiration_date - timezone.now().date()).days

        # Determine stock status
        if product.quantity == 0:
            stock_status = "RUPTURE"
            status_color = 'FFCDD2'  # Light red
        elif product.quantity <= 5:
            stock_status = "CRITIQUE"
            status_color = 'FFF3CD'  # Light yellow
        elif product.quantity <= 20:
            stock_status = "FAIBLE"
            status_color = 'D4EDDA'  # Light green
        else:
            stock_status = "NORMAL"
            status_color = 'FFFFFF'  # White

        # Clean description (remove HTML tags)
        description = clean_html(product.description) if product.description else "N/A"

        # Calculate total value
        total_value = float(product.quantity * product.price)

        # Data row
        row_data = [
            product.id,
            product.name,
            product.category.name if product.category else "N/A",
            description[:100] + "..." if len(description) > 100 else description,
            product.quantity,
            float(product.price),
            product.expiration_date.strftime('%d/%m/%Y') if product.expiration_date else "Non défini",
            remaining_days if remaining_days is not None else "N/A",
            stock_status,
            total_value,
            product.user.username if product.user else "N/A",
            timezone.now().strftime('%d/%m/%Y')
        ]

        # Add data to cells with formatting
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Apply status-based coloring
            if col_num == 9:  # Stock status column
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type='solid')
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Format currency columns
            elif col_num in [6, 10]:  # Price and Total Value columns
                cell.number_format = '#,##0.00 "€"'
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # Format quantity column
            elif col_num == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if value == 0:
                    cell.font = Font(name='Arial', size=10, bold=True, color='FF0000')

    # Auto-adjust column widths
    column_widths = [8, 25, 15, 30, 10, 12, 15, 12, 12, 15, 15, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add totals row
    total_row = len(products) + 5
    ws.merge_cells(f'A{total_row}:D{total_row}')
    total_label = ws[f'A{total_row}']
    total_label.value = "TOTAUX"
    total_label.font = Font(name='Arial', size=11, bold=True)
    total_label.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    total_label.alignment = Alignment(horizontal='center', vertical='center')

    # Total quantity
    total_qty_cell = ws.cell(row=total_row, column=5)
    total_qty_cell.value = f"=SUM(E4:E{total_row - 1})"
    total_qty_cell.font = Font(name='Arial', size=11, bold=True)
    total_qty_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')

    # Total value
    total_value_cell = ws.cell(row=total_row, column=10)
    total_value_cell.value = f"=SUM(J4:J{total_row - 1})"
    total_value_cell.font = Font(name='Arial', size=11, bold=True)
    total_value_cell.fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    total_value_cell.number_format = '#,##0.00 "€"'


def setup_summary_sheet(ws):
    """Setup summary dashboard sheet"""
    from .models import Product, Category
    from django.db.models import Sum, Count, Avg

    # Title
    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = "TABLEAU DE BORD - RÉSUMÉ EXÉCUTIF"
    title.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title.fill = PatternFill(start_color='2E86AB', end_color='2E86AB', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Get statistics
    products = Product.objects.all()
    total_products = products.count()
    total_quantity = products.aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_value = sum([p.quantity * p.price for p in products])
    avg_price = products.aggregate(Avg('price'))['price__avg'] or 0

    # Stock status counts
    out_of_stock = products.filter(quantity=0).count()
    low_stock = products.filter(quantity__lte=5, quantity__gt=0).count()
    normal_stock = products.filter(quantity__gt=20).count()

    # Expiring soon
    thirty_days_from_now = timezone.now().date() + timedelta(days=30)
    expiring_soon = products.filter(
        expiration_date__lte=thirty_days_from_now,
        expiration_date__gte=timezone.now().date()
    ).count()

    # Summary data
    summary_data = [
        ("📊 STATISTIQUES GÉNÉRALES", ""),
        ("Nombre total de produits", total_products),
        ("Quantité totale en stock", f"{total_quantity:,}"),
        ("Valeur totale du stock", f"{total_value:,.2f} €"),
        ("Prix moyen par produit", f"{avg_price:.2f} €"),
        ("", ""),
        ("⚠️ ALERTES STOCK", ""),
        ("Produits en rupture", out_of_stock),
        ("Stock critique (≤5)", low_stock),
        ("Stock normal (>20)", normal_stock),
        ("", ""),
        ("📅 DATES D'EXPIRATION", ""),
        ("Expirant sous 30 jours", expiring_soon),
        ("", ""),
        ("📈 PAR CATÉGORIE", ""),
    ]

    # Add category breakdown
    categories = Category.objects.annotate(
        product_count=Count('product'),
        total_qty=Sum('product__quantity')
    ).order_by('-product_count')

    for category in categories:
        summary_data.append((
            f"  • {category.name}",
            f"{category.product_count} produits ({category.total_qty or 0} unités)"
        ))

    # Add data to sheet
    for row_num, (label, value) in enumerate(summary_data, 3):
        label_cell = ws.cell(row=row_num, column=1)
        value_cell = ws.cell(row=row_num, column=2)

        label_cell.value = label
        value_cell.value = value

        if label.startswith(('📊', '⚠️', '📅', '📈')):
            # Section headers
            label_cell.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            label_cell.fill = PatternFill(start_color='A23B72', end_color='A23B72', fill_type='solid')
            value_cell.fill = PatternFill(start_color='A23B72', end_color='A23B72', fill_type='solid')
        elif label.startswith('  •'):
            # Sub-items
            label_cell.font = Font(name='Arial', size=10, italic=True)
            value_cell.font = Font(name='Arial', size=10)
        elif label and value:
            # Regular data
            label_cell.font = Font(name='Arial', size=11, bold=True)
            value_cell.font = Font(name='Arial', size=11)

    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25


def setup_alerts_sheet(ws):
    """Setup stock alerts sheet"""
    from .models import Product

    # Title
    ws.merge_cells('A1:F1')
    title = ws['A1']
    title.value = "🚨 ALERTES DE STOCK ET EXPIRATIONS"
    title.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    title.fill = PatternFill(start_color='DC3545', end_color='DC3545', fill_type='solid')
    title.alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    headers = ['Produit', 'Catégorie', 'Quantité', 'Date Expiration', 'Jours Restants', 'Type d\'Alerte']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='6C757D', end_color='6C757D', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Get products needing attention
    alerts = []

    # Out of stock
    out_of_stock = Product.objects.filter(quantity=0)
    for product in out_of_stock:
        alerts.append((product, "RUPTURE DE STOCK", 'FFCDD2'))

    # Low stock
    low_stock = Product.objects.filter(quantity__lte=5, quantity__gt=0)
    for product in low_stock:
        alerts.append((product, "STOCK CRITIQUE", 'FFF3CD'))

    # Expiring soon
    thirty_days = timezone.now().date() + timedelta(days=30)
    expiring = Product.objects.filter(
        expiration_date__lte=thirty_days,
        expiration_date__gte=timezone.now().date()
    )
    for product in expiring:
        days_left = (product.expiration_date - timezone.now().date()).days
        alert_type = "EXPIRE BIENTÔT" if days_left > 7 else "EXPIRE TRÈS BIENTÔT"
        color = 'FFF3CD' if days_left > 7 else 'FFCDD2'
        alerts.append((product, alert_type, color))

    # Add alert data
    for row_num, (product, alert_type, color) in enumerate(alerts, 4):
        remaining_days = None
        if product.expiration_date:
            remaining_days = (product.expiration_date - timezone.now().date()).days

        row_data = [
            product.name,
            product.category.name if product.category else "N/A",
            product.quantity,
            product.expiration_date.strftime('%d/%m/%Y') if product.expiration_date else "Non défini",
            remaining_days if remaining_days is not None else "N/A",
            alert_type
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = Font(name='Arial', size=10)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

            if col_num == 6:  # Alert type column
                cell.font = Font(name='Arial', size=10, bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    column_widths = [25, 15, 10, 15, 12, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def clean_html(text):
    """Remove HTML tags from text"""
    if not text:
        return ""
    clean = re.compile('<.*?>')
    return re.sub(clean, '', text).strip()

# ضع هنا مسار الخطوط في مشروعك

ARABIC_FONT_PATH = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansArabic-Regular.ttf")
ARABIC_BOLD_FONT_PATH = os.path.join(settings.BASE_DIR, "static", "fonts", "NotoSansArabic-Bold.ttf")


def process_arabic_text(text):
    if not text:
        return ""
    reshaped = arabic_reshaper.reshape(str(text))
    return get_display(reshaped)

def register_arabic_fonts():
    """
    Register Arabic fonts for ReportLab. Adjust paths above.
    Returns True if success, False otherwise.
    """
    try:
        pdfmetrics.registerFont(TTFont('Arabic', ARABIC_FONT_PATH))
        pdfmetrics.registerFont(TTFont('Arabic-Bold', ARABIC_BOLD_FONT_PATH))
        return True
    except Exception as e:
        # طباعة لغايات الديباغ فقط
        print("Arabic font registration failed:", e)
        return False


@login_required
def batch_pdf_arabic(request):
    """
    Generate a single Arabic PDF summarizing a batch of FicheStockSortie records.
    Expects list of fiche IDs in session under 'last_batch_ids' (as set by batch_sortie view).
    """
    from .models import FicheStockSortie  # local import to avoid circular issues

    fiche_ids = request.session.get('last_batch_ids', [])
    if not fiche_ids:
        return HttpResponse(process_arabic_text("لا يوجد بيانات لاستخراجها."), status=400)

    fiches = FicheStockSortie.objects.filter(id__in=fiche_ids).order_by('date')

    has_arabic = register_arabic_fonts()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    styles = getSampleStyleSheet()

    # Arabic styles (use registered font names if available)
    if has_arabic:
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontName='Arabic-Bold', fontSize=16, alignment=1)
        normal_r = ParagraphStyle('normal_r', parent=styles['Normal'], fontName='Arabic', fontSize=11, alignment=2)  # right-aligned
        header_r = ParagraphStyle('header_r', parent=styles['Normal'], fontName='Arabic-Bold', fontSize=12, alignment=2)
    else:
        title_style = ParagraphStyle('title', parent=styles['Heading1'], fontSize=16, alignment=1)
        normal_r = ParagraphStyle('normal_r', parent=styles['Normal'], fontSize=11, alignment=2)
        header_r = ParagraphStyle('header_r', parent=styles['Normal'], fontSize=12, alignment=2)

    # Document header (Arabic organization lines)
    elements.append(Paragraph(process_arabic_text("السّلطة الوطنية المستقلة للانتخابات"), title_style))
    elements.append(Paragraph(process_arabic_text("المندوبية الولائية لولاية بجاية"), title_style))
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(process_arabic_text("تقرير دفعة سندات إخراج المخزون"), header_r))
    elements.append(Spacer(1, 12))

    # Build table header (Arabic labels, keep content cells processed for RTL)
    table_data = [
        [
            process_arabic_text('التاريخ'),
            process_arabic_text('الوجهة'),
            process_arabic_text('الكمية'),
            process_arabic_text('اسم المنتج'),
            process_arabic_text('الرقم')
        ]
    ]

    # Fill rows: we'll prepare Arabic-processed strings for each cell where appropriate
    for f in fiches:
        date_str = f.date.strftime('%Y/%m/%d %H:%M')
        # process Arabic text for product name and destination (if present)
        prod_name = process_arabic_text(f.name_fiche.name)
        dest = process_arabic_text(f.destination or "غير محدد")
        num = process_arabic_text(f.number)
        qty = str(f.quantity)
        # For RTL table appearance, keep order of cells as in header but we will set alignment to RIGHT
        table_data.append([date_str, dest, qty, prod_name, num])

    # Build table with column widths (adjust as needed)
    col_widths = [100, 140, 60, 150, 100]  # in points (approx)
    table = Table(table_data, colWidths=col_widths, hAlign='RIGHT')

    # Table style: header background, right alignment, fonts use Arabic if available
    tbl_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0b5ed7')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arabic-Bold' if has_arabic else 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Arabic' if has_arabic else 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.beige]),
    ])
    table.setStyle(tbl_style)
    elements.append(table)
    elements.append(Spacer(1, 12))

    # Optionally, include per-fiche detailed block (observations) below
    for f in fiches:
        if f.observation:
            clean_obs = re.sub('<.*?>', '', f.observation)
            elements.append(Paragraph(process_arabic_text(f"ملاحظات حول السند {f.number}:"), header_r))
            elements.append(Paragraph(process_arabic_text(clean_obs), normal_r))
            elements.append(Spacer(1, 8))

    # Footer / signature area
    footer_table = Table([
        [process_arabic_text('توقيع المسؤول:'), ''],
        [process_arabic_text('التاريخ والتوقيع:'), '']
    ], colWidths=[200, 200], hAlign='RIGHT')
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,-1), 'Arabic' if has_arabic else 'Helvetica'),
        ('FONTSIZE', (0,0), (-1,-1), 11),
        ('LINEBELOW', (1,0), (1,0), 0.5, colors.black),
        ('LINEBELOW', (1,1), (1,1), 0.5, colors.black),
    ]))
    elements.append(Spacer(1, 20))
    elements.append(footer_table)

    # Build and return PDF
    doc.build(elements)
    pdf_value = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="batch_bon_livraison.pdf"'
    response.write(pdf_value)
    return response


def bon_livraison_arabic(request, fiche_id=None):
    """
    Generate Arabic PDF for a FicheStockSortie (or similar model).
    Requires: FicheStockSortie model imported where appropriate.
    """
    # استبدل الاستدعاء التالي بما يناسب مشروعك
    from .models import FicheStockSortie  # ضع import هنا لتجنب circular imports

    if fiche_id:
        fiche = get_object_or_404(FicheStockSortie, id=fiche_id)
    else:
        try:
            fiche = FicheStockSortie.objects.latest('date')
        except FicheStockSortie.DoesNotExist:
            return HttpResponse("لم يتم العثور على سند تسليم", status=404)

    has_arabic = register_arabic_fonts()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    elements = []
    styles = getSampleStyleSheet()

    # إنشاء أنماط عربية
    if has_arabic:
        arabic_style = ParagraphStyle('Arabic', parent=styles['Normal'], fontName='Arabic', fontSize=12, alignment=2)
        arabic_bold = ParagraphStyle('ArabicBold', parent=styles['Heading2'], fontName='Arabic-Bold', fontSize=14, alignment=1)
    else:
        arabic_style = ParagraphStyle('Arabic', parent=styles['Normal'], fontSize=12, alignment=2)
        arabic_bold = ParagraphStyle('ArabicBold', parent=styles['Heading2'], fontSize=14, alignment=1)

    # Header / Title
    elements.append(Paragraph(process_arabic_text("السّلطة الوطنية المستقلة للانتخابات"), arabic_bold))
    elements.append(Paragraph(process_arabic_text("المندوبية الولائية لولاية بجاية"), arabic_bold))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(process_arabic_text(f"سند تسليم رقم {fiche.number}"), arabic_bold))
    elements.append(Spacer(1, 16))

    # Date / meta table (Right label, Left value)
    date_info = [
        [fiche.date.strftime('%Y/%m/%d - %H:%M'), process_arabic_text('تاريخ الإنشاء:')],
        [fiche.user.username if fiche.user else 'غير محدد', process_arabic_text('المستخدم:')],
        [process_arabic_text(fiche.destination or ''), process_arabic_text('الوجهة:')],
    ]
    date_table = Table(date_info, colWidths=[2.2 * inch, 3.8 * inch])
    date_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (0, -1), 'Arabic-Bold' if has_arabic else 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Arabic' if has_arabic else 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey)
    ]))
    elements.append(date_table)
    elements.append(Spacer(1, 16))

    # Products table header + data
    product_details = [
        [
            process_arabic_text('اسم المنتج'),
            process_arabic_text('الكمية المخرجة'),
            process_arabic_text('المخزون المتبقي'),
            process_arabic_text('السعر')
        ],
        [
            process_arabic_text(fiche.name_fiche.name),
            str(fiche.quantity),
            str(fiche.name_fiche.quantity),
            f"{getattr(fiche.name_fiche, 'price', 'غير محدد')}"
        ]
    ]
    product_table = Table(product_details, colWidths=[2.5*inch, 1*inch, 1*inch, 1.2*inch])
    product_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0b5ed7')),  # header color
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Arabic-Bold' if has_arabic else 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Arabic' if has_arabic else 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.beige]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black)
    ]))
    elements.append(product_table)
    elements.append(Spacer(1, 12))

    # Observations
    if fiche.observation:
        import re
        clean_obs = re.sub('<.*?>', '', fiche.observation)
        elements.append(Paragraph(process_arabic_text("ملاحظات"), arabic_bold))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(process_arabic_text(clean_obs), arabic_style))
        elements.append(Spacer(1, 12))

    # Footer (signature)
    footer_data = [
        [process_arabic_text('توقيع المسؤول:'), ''],
        [process_arabic_text('التاريخ والتوقيع:'), '']
    ]
    footer_table = Table(footer_data, colWidths=[2.2*inch, 3.8*inch])
    footer_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic' if has_arabic else 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('LINEBELOW', (1, 0), (1, 0), 0.5, colors.black),
        ('LINEBELOW', (1, 1), (1, 1), 0.5, colors.black),
    ]))
    elements.append(Spacer(1, 20))
    elements.append(footer_table)

    # Build document
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="bon_livraison_{fiche.number}.pdf"'
    response.write(pdf)
    return response
